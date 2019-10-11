# stdlib imports
from xml.dom import minidom
import os.path
import math
import string
from functools import partial
import argparse

# third party imports
import pandas as pd
from shapely.geometry import shape as sShape, Point, MultiPolygon
import fiona
from obspy.clients.fdsn import Client
from impactutils.time.ancient_time import HistoricTime
from openpyxl import load_workbook
import pkg_resources
import pyproj
import numpy as np
from shapely.ops import transform
import requests

# local imports
from libcomcat.exceptions import ConnectionError
from libcomcat import __version__ as libversion

# use this to set the user agent for each request, giving us a way
# to distinguish libcomcat requests from other browser requests
HEADERS = {'User-Agent': 'libcomcat v%s' % libversion}

# constants
CATALOG_SEARCH_TEMPLATE = 'https://earthquake.usgs.gov/fdsnws/event/1/catalogs'
CONTRIBUTORS_SEARCH_TEMPLATE = ('https://earthquake.usgs.gov/fdsnws/event/'
                                '1/contributors')
TIMEOUT = 60
TIMEFMT1 = '%Y-%m-%dT%H:%M:%S'
TIMEFMT2 = '%Y-%m-%dT%H:%M:%S.%f'
DATEFMT = '%Y-%m-%d'
COUNTRYFILE = 'ne_10m_admin_0_countries.shp'

# where is the PAGER fatality model found?
FATALITY_URL = ('https://raw.githubusercontent.com/usgs/pager/master/'
                'losspager/data/fatality.xml')
ECONOMIC_URL = ('https://raw.githubusercontent.com/usgs/pager/master/'
                'losspager/data/economy.xml')

COUNTRIES_SHP = 'ne_50m_admin_0_countries.shp'
BUFFER_DISTANCE_KM = 100
KM_PER_DEGREE = 119.191

TIMEOUT = 60  # how long should we wait for a response from ComCat?


class CombinedFormatter(argparse.ArgumentDefaultsHelpFormatter,
                        argparse.RawTextHelpFormatter,
                        argparse.RawDescriptionHelpFormatter):
    pass


def get_mag_src(mag):
    """Try to find the best magnitude source from a Magnitude object.

    Note: This can be difficult, as there is a great deal of variance
    in how magnitude information is submitted in QuakeML within ComCat.

    Args:
        mag (obspy Magnitude): Magnitude object from obspy.
    Returns:
        str: String indicating the most likely source of the
             magnitude solution.

    """
    c1 = mag.creation_info is not None
    if c1:
        c2 = mag.creation_info.agency_id is not None
    else:
        c2 = False
    if c2:
        magsrc = mag.creation_info.agency_id.lower()
    else:
        has_gcmt = mag.resource_id.id.lower().find('gcmt') > -1
        has_at = mag.resource_id.id.lower().find('at') > -1
        has_pt = mag.resource_id.id.lower().find('pt') > -1
        has_ak = (mag.resource_id.id.lower().find('ak') > -1 or
                  mag.resource_id.id.lower().find('alaska') > -1)
        has_pr = mag.resource_id.id.lower().find('pr') > -1
        has_dup = mag.resource_id.id.lower().find('duputel') > -1
        has_us = mag.resource_id.id.lower().find('us') > -1
        if has_gcmt:
            magsrc = 'gcmt'
        elif has_dup:
            magsrc = 'duputel'
        elif has_at:
            magsrc = 'at'
        elif has_pt:
            magsrc = 'pt'
        elif has_ak:
            magsrc = 'ak'
        elif has_pr:
            magsrc = 'pr'
        elif has_us:
            magsrc = 'us'
        else:
            magsrc = 'unknown'

    return magsrc


def read_phases(filename):
    """Read a phase file CSV or Excel file into data structures.

    Args:
        filename (str): String file name of a CSV or Excel file
            created by getphases program.
    Returns:
        tuple:
            header_dict - Dictionary containing header data from top of file.
            dataframe - Pandas dataframe containing phase data.
    """
    if not os.path.isfile(filename):
        raise FileNotFoundError('Filename %s does not exist.' % filename)
    header_dict = {}
    if filename.endswith('xlsx'):
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.active
        key = ''
        rowidx = 1

        while key != 'Channel':
            key = ws['A%i' % rowidx].value
            if not key.startswith('#'):
                value = ws['B%i' % rowidx].value
                header_dict[key] = value
            rowidx += 1
        wb.close()
        dataframe = pd.read_excel(filename, skiprows=rowidx - 2)
    elif filename.endswith('csv'):
        f = open(filename, 'rt')
        tline = f.readline()
        rowidx = 0
        while tline.startswith('#'):
            if not tline.startswith('#%'):
                line = tline.replace('#', '').strip()
                key, value = line.split('=')
                key = key.strip()
                value = value.strip()
                header_dict[key] = value
            rowidx += 1
            tline = f.readline()
        f.close()
        dataframe = pd.read_csv(filename, skiprows=rowidx)
    else:
        f, ext = os.path.splitext(filename)
        raise Exception('Filenames with extension %s are not supported.' % ext)
    return (header_dict, dataframe)


def makedict(dictstring):
    try:
        parts = dictstring.split(':')
        key = parts[0]
        value = parts[1]
        return {key: value}
    except Exception:
        raise Exception(
            'Could not create a single key dictionary out of %s' % dictstring)


def maketime(timestring):
    outtime = None
    try:
        outtime = HistoricTime.strptime(timestring, TIMEFMT1)
    except Exception:
        try:
            outtime = HistoricTime.strptime(timestring, TIMEFMT2)
        except Exception:
            try:
                outtime = HistoricTime.strptime(timestring, DATEFMT)
            except Exception:
                raise Exception(
                    'Could not parse time or date from %s' % timestring)
    return outtime


def get_catalogs():
    """Get the list of catalogs available in ComCat.

    Returns:
        list: Catalogs available in ComCat (see the catalog
            parameter in search() method.)
    """
    try:
        request = requests.get(CATALOG_SEARCH_TEMPLATE, timeout=TIMEOUT)
        data = request.text
    except Exception as e:
        fmt = 'Could not connect to url %s. Error: "%s"'
        raise ConnectionError(fmt % (CATALOG_SEARCH_TEMPLATE, str(e)))

    root = minidom.parseString(data)
    catalogs = root.getElementsByTagName('Catalog')
    catlist = []
    for catalog in catalogs:
        catlist.append(catalog.firstChild.data)
    root.unlink()
    return catlist


def get_contributors():
    """Get the list of contributors available in ComCat.

    Returns:
        list: Contributors available in ComCat (see the contributor
            parameter in search() method.)
    """
    try:
        request = requests.get(CONTRIBUTORS_SEARCH_TEMPLATE, timeout=TIMEOUT)
        data = request.text
    except Exception as e:
        fmt = 'Could not connect to url %s. Error: "%s"'
        raise ConnectionError(fmt % (CONTRIBUTORS_SEARCH_TEMPLATE, str(e)))
    root = minidom.parseString(data)
    contributors = root.getElementsByTagName('Contributor')
    conlist = []
    for contributor in contributors:
        conlist.append(contributor.firstChild.data)
    root.unlink()
    return conlist


def check_ccode(ccode):
    """Ensure three letter country code is valid and contained in country bounds.

    Args:
        ccode (str): Three letter valid ISO 3166 country code.
    Returns:
        bool: True if valid country code found in bounds file, False otherwise.
    """
    ccode = ccode.upper()
    datapath = os.path.join('data', COUNTRIES_SHP)
    shpfile = pkg_resources.resource_filename('libcomcat', datapath)
    ccodes = []
    with fiona.open(shpfile, 'r') as shapes:
        for shape in shapes:
            isocode = shape['properties']['ADM0_A3']
            ccodes.append(isocode)
    if ccode not in ccodes:
        return False
    return True


def get_country_bounds(ccode, buffer_km=BUFFER_DISTANCE_KM):
    """Get list of country bounds (one for each sub-polygon in country polygon.)

    Args:
        ccode (str): Three letter ISO 3166 country code.
        buffer_km (int): Buffer distance around country boundary.

    Returns:
        list: List of 4-element tuples (xmin, xmax, ymin, ymax)

    """
    xmin = xmax = ymin = ymax = None
    ccode = ccode.upper()
    datapath = os.path.join('data', COUNTRIES_SHP)
    shpfile = pkg_resources.resource_filename('libcomcat', datapath)
    bounds = []
    with fiona.open(shpfile, 'r') as shapes:
        for shape in shapes:
            if shape['properties']['ADM0_A3'] == ccode:
                country = sShape(shape['geometry'])
                if isinstance(country, MultiPolygon):
                    for polygon in country:
                        xmin, ymin, xmax, ymax = _buffer(
                            polygon.bounds, buffer_km)
                        bounds.append((xmin, xmax, ymin, ymax))
                else:
                    xmin, ymin, xmax, ymax = _buffer(country.bounds, buffer_km)
                    bounds.append((xmin, xmax, ymin, ymax))
                break

    return bounds


def _buffer(bounds, buffer_km):
    xmin, ymin, xmax, ymax = bounds
    km2deg = (1 / KM_PER_DEGREE)
    ymin -= buffer_km * km2deg
    ymax += buffer_km * km2deg
    yav = (ymin + ymax) / 2
    xmin -= buffer_km * km2deg * np.cos(np.radians(yav))
    xmax += buffer_km * km2deg * np.cos(np.radians(yav))
    return (xmin, ymin, xmax, ymax)


def _get_country_shape(ccode):
    datapath = os.path.join('data', COUNTRIES_SHP)
    shpfile = pkg_resources.resource_filename('libcomcat', datapath)
    country = None
    with fiona.open(shpfile, 'r') as shapes:
        for shape in shapes:
            if shape['properties']['ADM0_A3'] == ccode:
                country = sShape(shape['geometry'])

    return country


def _get_utm_proj(lat, lon):
    zone = str((math.floor((lon + 180) / 6) % 60) + 1)
    alphabet = string.ascii_uppercase
    alphabet = alphabet.replace('I', '')
    alphabet = alphabet.replace('O', '')
    alphabet = alphabet[2:-2]
    if lat < -80:
        band = 'C'
    elif lat > 84:
        band = 'X'
    else:
        band_starts = np.arange(-80, 80, 8)
        # band_ends = np.append(np.arange(-72, 80, 8), [84])
        dstarts = lat - band_starts
        sidx = np.where(dstarts >= 0)[0].max()
        band = alphabet[sidx]
    fmt = ("+proj=utm +zone=%s%s, %s +ellps=WGS84 "
           "+datum=WGS84 +units=m +no_defs")
    south = ''
    if lat < 0:
        south = '+south'
    tpl = (zone, band, south)
    proj = pyproj.Proj(fmt % tpl)
    return proj


def _get_pshape(polygon, buffer_km):
    bounds = polygon.bounds  # xmin, ymin, xmax, ymax
    dlon = bounds[2] - bounds[0]
    if dlon < 0:
        dlon = bounds[2] + 360 - bounds[0]
    center_lon = bounds[0] + dlon / 2
    if center_lon > 180:
        center_lon -= 360
    center_lat = (bounds[1] + bounds[3]) / 2
    utmproj = _get_utm_proj(center_lat, center_lon)
    project = partial(
        pyproj.transform,
        pyproj.Proj(init='epsg:4326'),
        utmproj)

    pshape = transform(project, polygon)
    pshape = pshape.buffer(buffer_km * 1000)
    return (pshape, utmproj)


def filter_by_country(df, ccode, buffer_km=BUFFER_DISTANCE_KM):
    """Filter earthquake dataframe by country code.

    Args:
        df (DataFrame): pandas Dataframe with columns (latitude,longitude).
        ccode (str): Three letter ISO 3166 country code.
        buffer_km (int): Buffer distance around country boundary.

    Returns:
        DataFrame: Filtered dataframe.
    """
    pshapes = []
    shape = _get_country_shape(ccode)
    if isinstance(shape, MultiPolygon):
        for polygon in shape:
            pshape, utmproj = _get_pshape(polygon, buffer_km)
            pshapes.append((pshape, utmproj))
    else:
        pshape, utmproj = _get_pshape(shape, buffer_km)
        pshapes.append((pshape, utmproj))

    df2 = pd.DataFrame(columns=df.columns)
    for idx, row in df.iterrows():
        lat = row['latitude']
        lon = row['longitude']
        point_inside = False
        for pshape, utmproj in pshapes:
            x, y = utmproj(lon, lat)
            pxy = Point(x, y)
            if pshape.contains(pxy):
                point_inside = True
                break
            if point_inside:
                break
        if point_inside:
            df2 = df2.append(row)

    return df2
