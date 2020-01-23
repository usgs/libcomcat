#!/usr/bin/env python
import argparse
import sys
import os.path
from datetime import timedelta
import textwrap

# third party imports
import pandas as pd
from openpyxl import load_workbook, styles
from obspy.geodetics.base import gps2dist_azimuth
import numpy as np

# local imports
import libcomcat
from libcomcat.search import search, get_event_by_id
from libcomcat.dataframes import (get_history_data_frame, split_history_frame,
                                  PRODUCTS, TIMEFMT, PRODUCT_COLUMNS)
from libcomcat.logging import setup_logger

DISPLAY_TIME_FMT = '%Y-%m-%d %H:%M:%S'


class MyFormatter(argparse.RawTextHelpFormatter,
                  argparse.ArgumentDefaultsHelpFormatter):
    pass


COLORS = {'dyfi': 'FADBD8',
          'finite-fault': 'CD6155',
          'focal-mechanism': 'EC7063',
          'ground-failure': 'D1F2EB',
          'losspager': 'FCF3CF',
          'moment-tensor': 'F7DC6F',
          'oaf': '7FB3D5',
          'origin': 'C39BD3',
          'phase-data': 'E59866',
          'shakemap': '58D68D',
          'default': '99A3A4'}


def get_parser():
    desc = '''Print out ComCat event history.

    This program summarizes the history of an event through the
    products found in ComCat. The products that can be described are:
    'dyfi', 'finite-fault', 'focal-mechanism', 'ground-failure',
    'losspager', 'moment-tensor', 'oaf', 'origin', 'phase-data',
    and 'shakemap'.

    In the summary output (see below), all products will have the following
    columns:
    - Product:
        One of supported products (see above)
    - Authoritative Event ID:
        Authoritative ComCat event ID, mostly only useful when using -r flag.
    - Code:
        Event Source + Code, mostly only useful when using
        -r flag with geteventhist.
    - Associated:
        Boolean indicating whether this product is associated
        with authoritative event. (only appears when using -r flag).
    - Product Source:
        Network that contributed the product.
    - Product Version:
        Either ordinal number created by sorting products from a
        given source, or a version property set by the creator
        of the product.
    - Update Time:
        Time the product was sent, set either by PDL client or
        by the person or software that created the product
        (set as a property.)
    - Elapsed (min):
        Elapsed time in minutes between the update time and
        the *authoritative* origin time.
    - URL:
        The most representative URL for that *version* of the
        given product.
    - Description:
        Varies depending on the product, but all description
        fields are delineated first by a vertical pipe "|",
        and key/value pairs in each field are delineated
        by a hash "#". This is so that the --split option
        can parse the description column into many
        columns.

    To get one summary spreadsheet in Excel format listing all products
    for the M7.1 event from the 2019 Ridgecrest Sequence:

    geteventhist ci38457511 -d ~/tmp/ridgecrest -f excel

    To get one summary spreadhsheet as above, *excluding* DYFI products:

    geteventhist ci38457511 -d ~/tmp/ridgecrest -f excel -x dyfi

    To split out the "description" column into separate columns, and
    the products into separate spreadsheets:

    geteventhist ci38457511 -d ~/tmp/ridgecrest -f excel --split

    To retrieve summary information for only origins and shakemaps:

    geteventhist ci38457511 -d ~/tmp/ridgecrest -f excel -p origin shakemap

    To retrieve information for only origins and shakemaps, and split them
    into separate spreadsheets:

    geteventhist ci38457511 -d ~/tmp/ridgecrest -f excel -p origin shakemap --split

    To print one product table (say, origins) to stdout in HTML format:

    geteventhist ci38996632  -p origin --web --split > ~/test.html
    '''
    parser = argparse.ArgumentParser(description=desc,
                                     formatter_class=MyFormatter)
    # positional arguments
    parser.add_argument('eventid',
                        metavar='EVENTID', help='ComCat event ID.')

    # optional arguments
    ohelp = '''Directory where files are stored.'''
    parser.add_argument('-d', '--outdir', help=ohelp,
                        default=os.path.expanduser('~'))
    parser.add_argument('--exclude-products',
                        help='COMCAT products to be excluded from the spreadsheet.',
                        nargs='*', default=[])
    parser.add_argument('-f', '--format', help="Output format. Options include 'csv', 'tab', and 'excel'. Default is 'csv'.",
                        choices=['excel', 'csv', 'tab'],
                        default='csv', dest='format')
    loghelp = '''Send debugging, informational, warning and error messages to a file.
    '''
    parser.add_argument('--logfile', default='stderr', help=loghelp)
    levelhelp = '''Set the minimum logging level. The logging levels are(low to high):

     - debug: Debugging message will be printed, most likely for developers.
              Most verbose.
     - info: Only informational messages, warnings, and errors will be printed.
     - warning: Only warnings(i.e., could not retrieve information for a
                single event out of many) and errors will be printed.
     - error: Only errors will be printed, after which program will stop.
              Least verbose.
    '''
    parser.add_argument('--loglevel', default='info',
                        choices=['debug', 'info', 'warning', 'error'],
                        help=levelhelp)
    phelp = '''Limit to only the products specified. If no products are
    specified, all will be listed. See the full list of products here:
    See the full list here: https://usgs.github.io/pdl/userguide/products/index.html.
    '''
    parser.add_argument('-p', '--product-type', help=phelp,
                        nargs='*', default=[])
    rhelp = '''Search for other unassociated earthquakes
    inside a search radius (km). (Requires use of -w.)
    '''
    parser.add_argument('-r', '--radius', help=rhelp, type=float)
    shelp = 'Split descriptions of single-product queries into separate columns.'
    parser.add_argument('--split',
                        help=shelp,
                        action='store_true',
                        default=False)
    parser.add_argument('--version', action='version',
                        version=libcomcat.__version__, help='Version of libcomcat.')
    parser.add_argument('--web', help='Print HTML tables to stdout.',
                        default=False, action='store_true')
    whelp = 'Limit by time window in seconds. (Requires use of -r.)'
    parser.add_argument('-w', '--window', type=float,
                        help=whelp)
    return parser


def _mod_tframe(event, tevent, tframe):
    newframe = pd.DataFrame(columns=tframe.columns)
    tframe['Authoritative Event ID'] = event.id
    tframe['Associated'] = False
    for idx, row in tframe.iterrows():
        if row['Product'] not in ['origin', 'phase-data']:
            newframe = newframe.append(row)
        parts = row['Description'].split('|')
        authlat = event.latitude
        authlon = event.longitude
        authtime = event.time
        olat = tevent.latitude
        olon = tevent.longitude
        otime = tevent.time
        dist_m, _, _ = gps2dist_azimuth(authlat, authlon, olat, olon)
        dist = dist_m / 1000.0
        tdiff = (otime - authtime).total_seconds()
        dstr = 'Distance from auth. origin(km)# % .1f' % dist
        tstr = 'Offset from auth. origin (sec)# %.1f' % tdiff
        newparts = parts[0:-2] + [dstr, tstr]
        row['Description'] = '|'.join(newparts)
        newframe = newframe.append(row)
    return newframe


def save_dataframe(outdir, format, event, dataframe, product=None):
    border = styles.Border(left=styles.Side(border_style=None,
                                            color='FFFFFF'),
                           right=styles.Side(border_style=None,
                                             color='FFFFFF'),
                           top=styles.Side(border_style=None,
                                           color='FFFFFF'),
                           bottom=styles.Side(border_style=None,
                                              color='FFFFFF'),
                           diagonal=styles.Side(border_style=None,
                                                color='FFFFFF'),
                           diagonal_direction=0,
                           outline=styles.Side(border_style=None,
                                               color='FFFFFF'),
                           vertical=styles.Side(border_style=None,
                                                color='FFFFFF'),
                           horizontal=styles.Side(border_style=None,
                                                  color='FFFFFF')
                           )

    if format == 'excel':
        if product is not None:
            outfile = os.path.join(outdir,
                                   event.id + '_' + product + '.xlsx')
        else:
            outfile = os.path.join(outdir, event.id + '.xlsx')
        dataframe.to_excel(outfile, index=False)
        wb = load_workbook(outfile)
        ws = wb.active
        ws.insert_rows(0, amount=6)
        ws.cell(1, 1, value='Event ID')
        ws.cell(1, 2, value=event.id)
        ws.cell(2, 1, value='Origin Time')
        ws.cell(2, 2, value=event.time.strftime(TIMEFMT))
        ws.cell(3, 1, value='Magnitude')
        ws.cell(3, 2, value=event.magnitude)
        ws.cell(4, 1, value='Latitude')
        ws.cell(4, 2, value=event.latitude)
        ws.cell(5, 1, value='Longitude')
        ws.cell(5, 2, value=event.longitude)
        ws.cell(6, 1, value='Depth')
        ws.cell(6, 2, value=event.depth)

        fills = {}
        for product, color in COLORS.items():
            my_color = styles.colors.Color(rgb=color)
            my_fill = styles.fills.PatternFill(patternType='solid',
                                               fgColor=my_color)
            fills[product] = my_fill

        # color rows by product type
        for row in ws.iter_rows(min_row=8, min_col=2, max_col=2):
            mycell = row[0]
            if mycell.value in fills:
                fill = fills[mycell.value]
            else:
                fill = fills['default']
            row_range = '%i:%i' % (mycell.row, mycell.row)
            for cell in ws[row_range]:
                cell.fill = fill
                # TODO - figure out why this doesn't do anything!
                cell.border = border

        wb.save(outfile)
    else:
        if product is not None:
            outfile = os.path.join(outdir,
                                   event.id + '_' + product + '.csv')
        else:
            outfile = os.path.join(outdir, event.id + '.csv')
        if format == 'tab':
            dataframe.to_csv(outfile, sep='\t', index=False)
        else:
            dataframe.to_csv(outfile, index=False)
        cdata = open(outfile, 'rt').read()
        with open(outfile, 'wt') as f:
            f.write('# Event ID: %s\n' % event.id)
            f.write('# Origin Time: %s\n' % event.time.strftime(TIMEFMT))
            f.write('# Magnitude: %s\n' % event.magnitude)
            f.write('# Latitude: %s\n' % event.latitude)
            f.write('# Longitude: %s\n' % event.longitude)
            f.write('# Depth: %s\n' % event.depth)
            f.write(cdata)

    return outfile


def web_print(event, dataframe):
    etable_fmt = '''
    <pre >
    Event ID: % s
    Origin Time: % s
    Magnitude: % .1f
    Latitude: % .4f
    Longitude: % .4f
    Depth: % .1f
    </pre>
    '''
    etable_tpl = (event.id,
                  event.time.strftime(TIMEFMT),
                  event.magnitude,
                  event.latitude,
                  event.longitude,
                  event.depth)
    etable = etable_fmt % etable_tpl
    print(textwrap.dedent(etable))
    print(dataframe.to_html(index=False, border=0, max_rows=None, max_cols=None))


def simplify_times(dataframe):
    # re-format all time columns to be like: 2019-01-01 17:34:16.1
    # first figure out all the time-like columns
    # df['date'] = pd.to_datetime(df["date"].dt.strftime('%Y-%m'))
    dtypes = dataframe.dtypes
    for idx, dtype in dtypes.iteritems():
        if np.issubdtype(dtype, np.datetime64):
            dataframe[idx] = pd.to_datetime(
                dataframe[idx].dt.strftime(DISPLAY_TIME_FMT))


def main():
    pd.set_option('display.width', 1000)
    pd.set_option('display.max_rows', 1000)
    pd.set_option('display.max_colwidth', -1)
    pd.set_option('display.max_columns', 1000)
    pd.set_option("display.colheader_justify", "left")
    parser = get_parser()
    args = parser.parse_args()

    setup_logger(args.logfile, args.loglevel)

    # make sure that input products are in the list of supported products
    if not set(args.product_type) <= set(PRODUCTS):
        unsupported = list(set(args.product_type) - set(PRODUCTS))
        fmt = 'The following event products are not supported: '
        print(fmt % (','.join(unsupported)))
        sys.exit(1)

    # make sure that excluded products are in the list of supported products
    if not set(args.exclude_products) <= set(PRODUCTS):
        unsupported = list(set(args.exclude_products) - set(PRODUCTS))
        fmt = ('The following event products you want to exclude '
               'are not supported: ')
        print(fmt % (','.join(unsupported)))
        sys.exit(1)

    # web output and directory output are mutually exclusive
    if args.outdir and args.web:
        if args.outdir != os.path.expanduser('~'):
            msg = '''The -d and --web options are mutually exclusive, meaning
            that you cannot choose to write files to a directory and print
            HTML output to the screen simultaneously. Please choose one of
            those two options and try again.
            '''
            print(msg)
            sys.exit(1)
        else:
            args.outdir = None
    if args.radius and not args.window:
        msg = '''To define a time and distance range, the radius and window
         options must both be set.
        '''
        print(msg)
        sys.exit(1)
    if args.window and not args.radius:
        msg = '''To define a time and distance range, the radius and window
         options must both be set.
        '''
        print(msg)
        sys.exit(1)
    if args.product_type:
        products = args.product_type
    else:
        products = PRODUCTS

    if args.exclude_products:
        products = set(products) - set(args.exclude_products)

    try:
        detail_event = get_event_by_id(args.eventid, includesuperseded=True)
        dataframe, event = get_history_data_frame(detail_event, products)
    except Exception as e:
        fmt = '''Failed to retrieve event history data for
        event % s. Error message is as follows. Exiting.
        "%s"
        '''
        tpl = (args.eventid, str(e))
        print(fmt % tpl)
        sys.exit(1)

    if args.radius:
        radius_km = args.radius
        radius_secs = args.window
        stime = event.time - timedelta(seconds=radius_secs)
        etime = event.time + timedelta(seconds=radius_secs)

        eventlist = search(starttime=stime,
                           endtime=etime,
                           latitude=event.latitude,
                           longitude=event.longitude,
                           maxradiuskm=radius_km)
        for tevent in eventlist:
            if tevent.id == event.id:
                continue
            detail = tevent.getDetailEvent(includesuperseded=True)
            tframe = get_history_data_frame(detail, products)
            newframe = _mod_tframe(event, tevent, tframe)
            dataframe = dataframe.append(newframe, ignore_index=True)

        # now re-sort by update time
        dataframe = dataframe.sort_values('Update Time')
        dataframe = dataframe[PRODUCT_COLUMNS]
    else:
        # since "Authoritative Event ID" and "Associated" columns are only applicable when
        # we're including other events in our results, drop those columns
        # if we're not doing that.
        drop_columns = ['Authoritative Event ID', 'Associated']
        dataframe = dataframe.drop(drop_columns, axis='columns')

    if args.outdir is not None and not os.path.isdir(args.outdir):
        os.makedirs(args.outdir)

    if args.split:
        df_products = dataframe['Product'].unique().tolist()
        available_products = set(df_products) & set(products)
        # TODO: Consider merging phase-data and origin products
        # somehow in this process
        for product in available_products:
            pframe = split_history_frame(dataframe, product=product)
            simplify_times(pframe)
            if args.web:
                web_print(event, pframe)
            else:
                outfile = save_dataframe(args.outdir, args.format, event,
                                         pframe, product=product)
                print('%i rows saved to %s' % (len(pframe), outfile))
        sys.exit(0)

    if args.outdir:
        outfile = save_dataframe(args.outdir, args.format, event,
                                 dataframe, product=None)

        print('%i rows saved to %s' % (len(dataframe), outfile))
    elif args.web:
        simplify_times(dataframe)
        web_print(event, dataframe)

    sys.exit(0)


if __name__ == '__main__':
    main()
