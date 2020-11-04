#!/usr/bin/env python
import argparse
import sys
import logging

# third party imports
import pandas as pd
import openpyxl
from openpyxl.styles import Font, colors

# local imports
import libcomcat
from libcomcat.search import search, get_event_by_id
from libcomcat.classes import SummaryEvent
from libcomcat.utils import maketime
from libcomcat.dataframes import get_pager_data_frame
from libcomcat.logging import setup_logger


HEADER = '''
This data represents the results of running the PAGER exposure and loss
algorithms on the output from ShakeMap.

Notes: "Total" in the country column indicates that the results in that row are
the sum of exposures/losses for all affected countries.

"predicted_fatalities" and "predicted_dollars" are the results of applying loss
models to the exposure data - note that these values are not guaranteed to
match the actual losses from the earthquake.
'''


def add_headers(filename, file_format):
    headers = HEADER.split('\n')
    headers = ['#' + h for h in headers]

    if file_format == 'csv':
        data = open(filename, 'rt').read()
        headertext = '\n'.join(headers) + '\n'
        data = headertext + data
        with open(filename, 'wt') as f:
            f.write(data)
    else:
        font = Font(color="FF0000", bold=True)
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.insert_rows(1, amount=len(headers))
        for cellidx in range(0, len(headers)):
            coordinate = 'A%i' % (cellidx + 1)
            ws[coordinate] = headers[cellidx].strip('#')
            cell = ws[coordinate]
            cell.font = font
            wb.save(filename)
            wb.close()


def get_parser():
    desc = '''Download PAGER exposure/loss results in line format (csv, tab, etc.).

    To download basic PAGER information (total exposure) for events around New
    Zealand from 2010 to the present in CSV format:

    %(prog)s nz_exposures.csv -f csv -s 2010-01-01 -m 5.5 9.9 -b 163.213 -178.945 -48.980 -32.324

    To download the same information in Excel format:

    %(prog)s nz_exposures.xlsx -f excel -s 2010-01-01 -m 5.5 9.9 -b 163.213 -178.945 -48.980 -32.324

    To add loss information (see notes below), you can use the -l flag:

    %(prog)s nz_exposures.xlsx -f excel -s 2010-01-01 -m 5.5 9.9 -b 163.213 -178.945 -48.980 -32.324 -l

    To add exposures on a per-country basis (see notes below), you can use the -c flag:

    %(prog)s nz_exposures.xlsx -f excel -s 2010-01-01 -m 5.5 9.9 -b 163.213 -178.945 -48.980 -32.324 -c

    NOTES:

    1) Any start or end time where only date is specified (YYYY-mm-dd) will
    be translated to the beginning of that day.  Thus, a start time of
    "2015-01-01" becomes "2015-01-01T:00:00:00" and an end time of "2015-01-02"
    becomes ""2015-01-02T:00:00:00".

    2) Older events may not have the predicted loss information in ComCat - in
    those cases, predicted losses and uncertainties will be filled in with NaN
    values.

    3) Older events may not have the per-country exposure information
    available in ComCat.

    4) Note that when specifying a search box that crosses the -180/180
    meridian, you simply specify longitudes as you would if you were
    not crossing that meridian (i.e., lonmin=179, lonmax=-179).  The
    program will resolve the discrepancy.

    5) The ComCat API has a returned event limit of 20,000.  Queries that
    exceed this ComCat limit ARE supported by this software, by
    breaking up one large request into a number of smaller ones.'''

    parser = argparse.ArgumentParser(
        description=desc, formatter_class=argparse.RawDescriptionHelpFormatter)
    # positional arguments
    parser.add_argument('filename',
                        metavar='FILENAME', help='Output file name.')
    # optional arguments
    versionhelp = 'Retrieve information from all versions of PAGER.'
    parser.add_argument('-a', '--all', help=versionhelp,
                        action='store_true',
                        default=False)
    helpstr = ('Bounds to constrain event search '
               '[lonmin lonmax latmin latmax].')
    parser.add_argument('-b', '--bounds',
                        metavar=('lonmin', 'lonmax', 'latmin', 'latmax'),
                        dest='bounds', type=float, nargs=4,
                        help=helpstr)
    helpstr = ('End time for search (defaults to current date/time). '
               'YYYY-mm-dd, YYYY-mm-ddTHH:MM:SS, or YYYY-mm-ddTHH:MM:SS.s.')
    parser.add_argument('-e', '--end-time', dest='endTime', type=maketime,
                        help=helpstr)
    parser.add_argument('-f', '--format', dest='format',
                        choices=['csv', 'tab', 'excel'], default='csv',
                        metavar='FORMAT', help="Output format (csv, tab, or excel). Default is ‘csv’.")
    countryhelp = ('Retrieve information from all countries affected '
                   'by the earthquake.')
    parser.add_argument('--get-countries', help=countryhelp,
                        action='store_true',
                        default=False)
    losshelp = 'Retrieve fatalities and economic losses.'
    parser.add_argument('--get-losses', help=losshelp,
                        action='store_true',
                        default=False)
    helpstr = ('Specify a different comcat *search* host than '
               'earthquake.usgs.gov.')
    parser.add_argument('--host',
                        help=helpstr)
    versionhelp = 'Retrieve information from a single PAGER event, using ComCat event ID.'
    parser.add_argument('-i', '--eventid', help=versionhelp,
                        metavar='EVENTID')
    loghelp = '''Send debugging, informational, warning and error messages to a file.
    '''
    parser.add_argument('--logfile', default='stderr', help=loghelp)
    levelhelp = '''Set the minimum logging level. The logging levels are (low to high):

     - debug: Debugging message will be printed, most likely for developers.
              Most verbose.
     - info: Only informational messages, warnings, and errors will be printed.
     - warning: Only warnings (i.e., could not retrieve information for a
                single event out of many) and errors will be printed.
     - error: Only errors will be printed, after which program will stop.
              Least verbose.
    '''
    parser.add_argument('--loglevel', default='info',
                        choices=['debug', 'info', 'warning', 'error'],
                        help=levelhelp)
    helpstr = 'Minimum and maximum (authoritative) magnitude to restrict search.'
    parser.add_argument('-m', '--mag-range', metavar=('minmag', 'maxmag'),
                        dest='magRange', type=float, nargs=2,
                        help=helpstr)
    helpstr = ('Search radius in kilometers (radius and bounding options are '
               'mutually exclusive). The latitude and longitude for the '
               'search should be specified before the radius.')
    parser.add_argument('-r', '--radius', dest='radius',
                        metavar=('lat', 'lon', 'rmax'),
                        type=float, nargs=3,
                        help=helpstr)
    helpstr = ('Start time for search (defaults to ~30 days ago). '
               'YYYY-mm-dd, YYYY-mm-ddTHH:MM:SS, or YYYY-mm-ddTHH:MM:SS.s.')
    parser.add_argument('-s', '--start-time', dest='startTime', type=maketime,
                        help=helpstr)
    helpstr = ('Limit to events after specified time. YYYY-mm-dd or '
               'YYYY-mm-ddTHH:MM:SS.')
    parser.add_argument('-t', '--time-after', dest='after', type=maketime,
                        help=helpstr)
    parser.add_argument('--version', action='version',
                        version=libcomcat.__version__, help='Version of libcomcat.')
    return parser


def main():
    parser = get_parser()
    args = parser.parse_args()

    setup_logger(args.logfile, args.loglevel)

    latitude = None
    longitude = None
    radiuskm = None
    lonmin = latmin = lonmax = latmax = None
    if args.radius:
        latitude = args.radius[0]
        longitude = args.radius[1]
        radiuskm = args.radius[2]

    if args.bounds:
        lonmin, lonmax, latmin, latmax = args.bounds
        # fix longitude bounds when crossing dateline
        if lonmin > lonmax and lonmax >= -180:
            lonmin -= 360
    else:
        lonmin, lonmax, latmin, latmax = None, None, None, None

    minmag = 0.0
    maxmag = 9.9
    if args.magRange:
        minmag = args.magRange[0]
        maxmag = args.magRange[1]

    if args.bounds and args.radius:
        print('Please specify either a bounding box OR radius search.')
        sys.exit(1)

    if args.eventid:
        event = get_event_by_id(args.eventid,
                                includesuperseded=args.all)
        events = [event]
    else:
        events = search(starttime=args.startTime,
                        endtime=args.endTime,
                        updatedafter=args.after,
                        minlatitude=latmin,
                        maxlatitude=latmax,
                        minlongitude=lonmin,
                        maxlongitude=lonmax,
                        latitude=latitude,
                        longitude=longitude,
                        maxradiuskm=radiuskm,
                        maxmagnitude=maxmag,
                        minmagnitude=minmag,
                        producttype='losspager',
                        host=args.host)

    if not len(events):
        print('No events found matching your search criteria. Exiting.')
        sys.exit(0)

    dataframe = None
    nevents = len(events)
    i = 1
    for event in events:
        logging.debug('Processing event %s (%i of %i).\n' %
                      (event.id, i, nevents))

        if isinstance(event, SummaryEvent):
            detail = event.getDetailEvent(includesuperseded=args.all)
        else:
            detail = event
        df = get_pager_data_frame(detail, get_losses=args.get_losses,
                                  get_country_exposures=args.get_countries,
                                  get_all_versions=args.all)
        if dataframe is None:
            dataframe = df
        else:
            dataframe = pd.concat([dataframe, df])

    if dataframe is not None:
        logging.debug('Created table...saving %i records to %s.\n' %
                      (len(dataframe), args.filename))
        if args.format == 'excel':
            dataframe.to_excel(args.filename, index=False)
        elif args.format == 'tab':
            dataframe.to_csv(args.filename, sep='\t', index=False)
        else:
            dataframe.to_csv(args.filename, index=False, chunksize=1000)

        add_headers(args.filename, args.format)
        print('%i records saved to %s.' % (len(dataframe), args.filename))
    else:
        sys.stderr.write('No Pager products found for requested event(s)\n')
    sys.exit(0)


if __name__ == '__main__':
    main()
